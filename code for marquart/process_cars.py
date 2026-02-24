import pandas as pd
from scipy.stats import percentileofscore
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import re

def clean_price(val):
    """Convert '$46,999 ' -> 46999.0"""
    if pd.isna(val):
        return None
    return float(re.sub(r'[^\d.]', '', str(val)))

def clean_odometer(val):
    """Convert '15,189 mi' -> 15189.0"""
    if pd.isna(val):
        return None
    return float(re.sub(r'[^\d.]', '', str(val)))

def process_file(input_path, output_path):
    # Read CSV
    df = pd.read_csv(input_path, on_bad_lines='skip')

    # Drop completely empty rows
    df = df.dropna(how='all').reset_index(drop=True)

    # Clean Price and Odometer columns
    df['Price_clean'] = df['Price'].apply(clean_price)
    df['Odometer_clean'] = df['Odometer'].apply(clean_odometer)

    # Drop rows missing price or odometer
    df = df.dropna(subset=['Price_clean', 'Odometer_clean']).reset_index(drop=True)

    prices = df['Price_clean'].tolist()
    ages = df['Odometer_clean'].tolist()

    # PERCENTRANK × -1 (higher score = better = cheaper/lower mileage)
    df['Relative Price'] = df['Price_clean'].apply(
        lambda p: round(percentileofscore(prices, p, kind='rank') / 100 * -1, 4)
    )
    df['Relative Age'] = df['Odometer_clean'].apply(
        lambda a: round(percentileofscore(ages, a, kind='rank') / 100 * -1, 4)
    )
    df['Score'] = (df['Relative Price'] + df['Relative Age']).round(4)

    # Rank: highest score = Rank 1
    df['Rank'] = df['Score'].rank(ascending=False, method='min').astype(int)

    # Sort by rank
    df = df.sort_values('Rank').reset_index(drop=True)

    # Drop helper columns
    df = df.drop(columns=['Price_clean', 'Odometer_clean'])

    # --- Write to styled Excel ---
    df.to_excel(output_path, index=False, engine='openpyxl')

    wb = load_workbook(output_path)
    ws = wb.active

    # Styles
    header_font = Font(name='Arial', bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill('solid', start_color='1F4E79')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

    rank1_fill = PatternFill('solid', start_color='C6EFCE')   # green - top 10
    rank2_fill = PatternFill('solid', start_color='FFEB9C')   # yellow - middle
    rank3_fill = PatternFill('solid', start_color='FFC7CE')   # red - bottom 10

    thin = Side(style='thin', color='CCCCCC')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    total_rows = ws.max_row - 1  # excluding header
    top10 = max(1, round(total_rows * 0.27))
    bottom10 = max(1, round(total_rows * 0.27))

    # Find column indices for key columns
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    rank_col = headers.index('Rank') + 1 if 'Rank' in headers else None

    # Style header row
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(1, col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = border

    ws.row_dimensions[1].height = 30

    # Style data rows
    for row in range(2, ws.max_row + 1):
        rank_val = ws.cell(row, rank_col).value if rank_col else None
        data_row = row - 1

        if rank_val is not None:
            if data_row <= top10:
                row_fill = rank1_fill
            elif data_row > total_rows - bottom10:
                row_fill = rank3_fill
            else:
                row_fill = rank2_fill
        else:
            row_fill = None

        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row, col)
            cell.font = Font(name='Arial', size=10)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
            if row_fill:
                cell.fill = row_fill

    # Auto-size columns
    for col in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col)
        max_len = max(
            len(str(ws.cell(r, col).value or ''))
            for r in range(1, ws.max_column + 1)
        )
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 8), 35)

    # Freeze top row
    ws.freeze_panes = 'A2'

    wb.save(output_path)
    print(f"✅ Done! Saved to: {output_path}")
    return df


if __name__ == '__main__':
    import sys
    if len(sys.argv) < 3:
        print("Usage: python process_cars.py input.csv output.xlsx")
    else:
        process_file(sys.argv[1], sys.argv[2])
