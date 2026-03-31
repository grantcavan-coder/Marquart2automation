import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import re

def clean_price(val):
    if pd.isna(val):
        return None
    return float(re.sub(r'[^\d.]', '', str(val)))

def clean_odometer(val):
    if pd.isna(val):
        return None
    return float(re.sub(r'[^\d.]', '', str(val)))

def percentrank(values, x):
    """Replicates Excel's PERCENTRANK: fraction of values strictly less than x."""
    values = [v for v in values if v is not None]
    n = len(values)
    if n <= 1:
        return 0
    below = sum(1 for v in values if v < x)
    return below / (n - 1)

def process_file(input_path, output_path):
    if input_path.endswith('.xlsx'):
        df = pd.read_excel(input_path, engine='openpyxl')
    else:
        df = pd.read_csv(input_path, on_bad_lines='skip')
    
    df = df.astype({col: 'object' for col in df.select_dtypes(include='string').columns})

    df['Price_clean'] = df['Price'].apply(clean_price)
    df['Odometer_clean'] = df['Odometer'].apply(clean_odometer)
    df = df.dropna(subset=['Price_clean', 'Odometer_clean']).reset_index(drop=True)

    prices = df['Price_clean'].tolist()
    ages = df['Odometer_clean'].tolist()

    df['Relative Price'] = df['Price_clean'].apply(
        lambda p: round(percentrank(prices, p) * -1, 4)
    )
    df['Relative Age'] = df['Odometer_clean'].apply(
        lambda a: round(percentrank(ages, a) * -1, 4)
    )
    df['Score'] = (df['Relative Price'] + df['Relative Age']).round(4)
    df['Rank'] = df['Score'].rank(ascending=False, method='min').astype(int)
    df = df.sort_values('Rank').reset_index(drop=True)
    df = df.drop(columns=['Price_clean', 'Odometer_clean'])

    df.to_excel(output_path, index=False, engine='openpyxl')

    wb = load_workbook(output_path)
    ws = wb.active

    header_font = Font(name='Arial', bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill('solid', start_color='1F4E79')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

    rank1_fill = PatternFill('solid', start_color='C6EFCE')
    rank2_fill = PatternFill('solid', start_color='FFEB9C')
    rank3_fill = PatternFill('solid', start_color='FFC7CE')

    thin = Side(style='thin', color='CCCCCC')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    total_rows = ws.max_row - 1
    top_n = max(1, round(total_rows * 0.27))
    bottom_n = max(1, round(total_rows * 0.27))

    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    rank_col = headers.index('Rank') + 1 if 'Rank' in headers else None

    for col in range(1, ws.max_column + 1):
        cell = ws.cell(1, col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = border
    ws.row_dimensions[1].height = 30

    for row in range(2, ws.max_row + 1):
        data_row = row - 1
        if rank_col:
            if data_row <= top_n:
                row_fill = rank1_fill
            elif data_row > total_rows - bottom_n:
                row_fill = rank3_fill
            else:
                row_fill = rank2_fill
        else:
            row_fill = None

        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row, col)
            cell.font = Font(name='Arial', size=10)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
            if row_fill:
                cell.fill = row_fill

    for col in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col)
        max_len = max(
            len(str(ws.cell(r, col).value or ''))
            for r in range(1, min(ws.max_row + 1, 50))
        )
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 8), 35)

    ws.freeze_panes = 'A2'
    wb.save(output_path)
    return df


if __name__ == '__main__':
    import sys
    if len(sys.argv) < 3:
        print("Usage: python process_cars.py input.csv output.xlsx")
    else:
        process_file(sys.argv[1], sys.argv[2])


if __name__ == '__main__':
    import sys
    if len(sys.argv) < 3:
        print("Usage: python process_cars.py input.csv output.xlsx")
    else:
        process_file(sys.argv[1], sys.argv[2])
